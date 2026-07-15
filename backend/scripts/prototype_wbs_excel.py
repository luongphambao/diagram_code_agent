"""Phase-0 prototype: prove we can generate a BnK-format WBS .xlsx with live
formulas + variable rows BEFORE wiring any agent tools.

Builds a small but representative wbs dict (3 phases, sub-module groups, mixed
phase_types), clones the template, writes the workbook, then re-opens it and
checks every leaf/rollup formula references valid rows and that the Python-side
roll-up (wbs_effort) reconciles. Run:

    python backend/scripts/prototype_wbs_excel.py
"""

from __future__ import annotations

import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]            # backend/
sys.path.insert(0, str(ROOT / "src"))
REPO = ROOT.parent
TEMPLATE = REPO / "DATA" / "[BnK] Template - WBS.xlsx"
OUT = REPO / "DATA" / "SOLUTION_WBS" / "wbs_prototype.xlsx"

from domain.wbs.wbs_effort import derive_leaf_effort, rollup, delivery_grid  # noqa: E402
import domain.wbs.wbs_excel as wbs_excel  # noqa: E402


def leaf(name, *, be=0, fe=0, mobile=0, ai=0, ba=0, pt="development", desc="", remark=""):
    e = derive_leaf_effort(be=be, fe=fe, mobile=mobile, ai=ai, ba=ba, phase_type=pt)
    return {
        "name": name, "description": desc, "remark": remark, "phase_type": pt,
        "be": e["be"], "fe": e["fe"], "mobile": e["mobile"], "ai": e["ai"],
        "fe_mobile": round(e["fe"] + e["mobile"] + e["ai"], 4),
        "ba": e["ba"], "qc": e["qc"], "pm": e["pm"], "total": e["total"],
    }


def sample_wbs() -> dict:
    return {
        "project_info": {"name": "Demo Lending Platform", "project_code": "DLP"},
        "ratios": {"ba_on_dev": 0.10, "qc_on_dev": 0.30, "pm_on_total": 0.10},
        "phases": [
            {"code": "I", "name": "SET UP & INSTALLATION", "modules": [
                {"code": "I.A", "name": "Solution Design", "groups": [
                    {"name": None, "items": [
                        leaf("Database Design", be=1, pt="design", desc="Schema + DB cluster"),
                        leaf("System Design", be=5, pt="design"),
                        leaf("UI/UX Design", mobile=10, pt="uiux"),
                    ]},
                ]},
                {"code": "I.B", "name": "System Operation", "groups": [
                    {"name": None, "items": [
                        leaf("Deployment Setup", be=5, pt="design"),
                        leaf("Infrastructure Monitoring", be=2, pt="design"),
                    ]},
                ]},
            ]},
            {"code": "II", "name": "DEVELOPMENT", "modules": [
                {"code": "II.A", "name": "Web Portal", "groups": [
                    {"name": "Common Module", "items": [
                        leaf("Authentication & Authorization", be=0.5, fe=0.5),
                        leaf("Common Frame", fe=1.5),
                    ]},
                    {"name": "Account Module", "items": [
                        leaf("Registration", be=3.5, fe=5),
                        leaf("Dashboard", be=1.5, fe=1.5),
                    ]},
                ]},
                {"code": "II.B", "name": "Core Service", "groups": [
                    {"name": "Notification", "items": [
                        leaf("Notification Adapter", be=2),
                        leaf("Email Notification", be=1),
                        leaf("SMS Notification", be=1),
                    ]},
                ]},
            ]},
            {"code": "III", "name": "TESTING & DEPLOYMENT SUPPORT", "modules": [
                {"code": "III.A", "name": "Solution Qualification", "groups": [
                    {"name": None, "items": [
                        leaf("Fix SIT/UAT Issues", be=5, fe=2.5, pt="support",
                             remark="Room for client enhancement; no charge"),
                    ]},
                ]},
                {"code": "III.B", "name": "Deployment & Maintenance", "groups": [
                    {"name": None, "items": [
                        leaf("Production Deployment", be=1, pt="deployment"),
                        leaf("Post Go-live Support", be=5, pt="design"),
                    ]},
                ]},
            ]},
        ],
        "timeline": {"weeks": 16},
        "milestones": [],
    }


def collect_leaves(wbs):
    out = []
    for p in wbs["phases"]:
        for m in p["modules"]:
            for g in m["groups"]:
                out.extend(g["items"])
    return out


def main():
    wbs = sample_wbs()
    leaves = collect_leaves(wbs)
    roll = rollup(leaves)
    print(f"[python rollup] total_mandays={roll['total_mandays']} "
          f"MM={roll['total_manmonths']} by_role={roll['effort_by_role']}")
    print(f"[delivery grid] {delivery_grid(wbs['timeline']['weeks'])}")

    layout = wbs_excel.build_wbs_workbook(wbs, OUT, TEMPLATE)
    print(f"[built] {OUT}  modules={list(layout['module_rows'])}  "
          f"phase_rows={layout['phase_rows']}  last_row={layout['last_row']}")

    # Re-open and dump the 2. WBS sheet to eyeball formulas.
    from openpyxl import load_workbook
    ws = load_workbook(OUT, data_only=False)["2. WBS"]
    print("\n--- 2. WBS generated (B|C|D | F-total | G-BE H-FE I-RA J-QC K-PM) ---")
    print(f" D2(proj)={ws.cell(2,4).value!r}  B3(title)={ws.cell(3,2).value!r}")
    for r in range(6, layout["last_row"] + 1):
        vals = [ws.cell(r, c).value for c in (2, 3, 4, 6, 7, 8, 9, 10, 11)]
        if any(v is not None for v in vals):
            cells = " | ".join(str(v)[:26] for v in vals if v is not None)
            print(f" R{r}: {cells}")

    # sanity: every formula references an existing row <= last_row
    bad = []
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if isinstance(v, str) and v.startswith("="):
                import re
                for ref in re.findall(r"[A-Z]+(\d+)", v):
                    if int(ref) > layout["last_row"] + 1:
                        bad.append((cell.coordinate, v))
    print(f"\n[formula refs] {'OK — all in range' if not bad else f'BAD: {bad[:5]}'}")

    # --- mini formula evaluator: prove Excel WILL recompute to the Python rollup ---
    md = load_workbook(OUT, data_only=False)["4. Master Data"]
    md_vals = {"'4. Master Data'!$C$5": md["C5"].value,
               "'4. Master Data'!$C$7": md["C7"].value,
               "'4. Master Data'!$C$4": md["C4"].value}
    recomputed = _evaluate_wbs_sheet(ws, layout["last_row"], md_vals)
    gt = {col: recomputed.get(f"{col}6", 0.0) for col in "GHIJK"}
    excel_total = round(sum(gt.values()), 2)
    print(f"[excel recompute] grand total per role G..K={ {k: round(v,2) for k,v in gt.items()} } "
          f"sum={excel_total}")
    ok = abs(excel_total - roll["total_mandays"]) < 0.01
    print(f"[RECONCILE] excel_total={excel_total} vs python_rollup={roll['total_mandays']} "
          f"-> {'MATCH ✓' if ok else 'MISMATCH ✗'}")

    # --- Delivery Plan checks ---
    dly = layout["delivery"]
    wd = load_workbook(OUT, data_only=False)["3. Delivery Plan"]
    months_hdr = sum(1 for c in range(6, 6 + dly["weeks"]) if str(wd.cell(3, c).value or "").endswith("Month"))
    week_hdr = sum(1 for c in range(6, 6 + dly["weeks"]) if str(wd.cell(5, c).value or "").startswith("W"))
    print(f"[delivery 16wk] months={dly['months']} sprints={dly['sprints']} weeks={dly['weeks']} "
          f"| header month-cells={months_hdr} week-cells={week_hdr} "
          f"| schedule={[(m['code'], m['start_week'], m['end_week']) for m in dly['schedule']]}")
    overflow = [m['code'] for m in dly['schedule'] if m['end_week'] > dly['weeks']]
    print(f"[delivery overflow] {'none ✓' if not overflow else overflow}")

    # --- 8-month (≥32 week) test: ensure enough months ---
    big = sample_wbs(); big["timeline"]["weeks"] = 33
    OUT8 = OUT.with_name("wbs_prototype_8mo.xlsx")
    lay8 = wbs_excel.build_wbs_workbook(big, OUT8, TEMPLATE)
    d8 = lay8["delivery"]
    w8 = load_workbook(OUT8, data_only=False)["3. Delivery Plan"]
    m8 = sum(1 for c in range(6, 6 + d8["weeks"]) if str(w8.cell(3, c).value or "").endswith("Month"))
    print(f"[delivery 33wk] months={d8['months']} (expect 9) sprints={d8['sprints']} weeks={d8['weeks']} "
          f"month-header-cells={m8} -> {'ENOUGH MONTHS ✓' if d8['months']==9 and d8['weeks']==33 else 'WRONG ✗'}")
    print(f"\nOpen {OUT} / {OUT8} in Excel to confirm visually.")


def _evaluate_wbs_sheet(ws, last_row, md_vals):
    """Tiny evaluator for the generated 2. WBS sheet: resolves SUM ranges/lists and
    leaf ratio formulas (which reference Master Data) so we can confirm the live
    workbook recomputes to the same totals as the Python rollup — without Excel."""
    import re
    cache: dict[str, float] = {}

    def cell_val(coord):
        if coord in cache:
            return cache[coord]
        cache[coord] = 0.0  # guard against cycles
        col = re.match(r"([A-Z]+)(\d+)", coord).group(1)
        row = int(re.match(r"([A-Z]+)(\d+)", coord).group(2))
        v = ws[coord].value
        cache[coord] = _eval(v, col, row)
        return cache[coord]

    def _eval(v, col, row):
        if v is None:
            return 0.0
        if isinstance(v, (int, float)):
            return float(v)
        if not (isinstance(v, str) and v.startswith("=")):
            return 0.0
        expr = v[1:]
        for ref, val in md_vals.items():           # substitute Master Data constants
            expr = expr.replace(ref, str(val))
        # expand SUM(...)
        def sum_repl(m):
            inner = m.group(1)
            total = 0.0
            for part in inner.split(","):
                part = part.strip()
                if ":" in part:                     # rectangular range A1:C9
                    a, b = part.split(":")
                    cA = re.match(r"([A-Z]+)(\d+)", a); cB = re.match(r"([A-Z]+)(\d+)", b)
                    from openpyxl.utils import column_index_from_string, get_column_letter
                    c0 = column_index_from_string(cA.group(1)); c1 = column_index_from_string(cB.group(1))
                    for ci in range(min(c0, c1), max(c0, c1) + 1):
                        for rr in range(int(cA.group(2)), int(cB.group(2)) + 1):
                            total += cell_val(f"{get_column_letter(ci)}{rr}")
                elif re.match(r"^[A-Z]+\d+$", part):
                    total += cell_val(part)
                else:
                    try:
                        total += float(part)
                    except ValueError:
                        pass
            return str(total)
        prev = None
        while prev != expr:
            prev = expr
            expr = re.sub(r"SUM\(([^()]*)\)", sum_repl, expr)
        # now expr is arithmetic with cell refs left (e.g. "12.0*0.1") — resolve refs
        expr = re.sub(r"[A-Z]+\d+", lambda m: str(cell_val(m.group(0))), expr)
        expr = re.sub(r"CONCATENATE\([^)]*\)", "0", expr)
        try:
            return float(eval(expr, {"__builtins__": {}}, {}))
        except Exception:
            return 0.0

    out = {}
    for col in "GHIJK":
        out[f"{col}6"] = cell_val(f"{col}6")
    return out


if __name__ == "__main__":
    main()
