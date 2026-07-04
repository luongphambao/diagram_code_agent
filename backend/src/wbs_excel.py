"""Build a filled BnK-format WBS workbook from a normalized ``wbs`` dict.

This is the hard, high-risk part of the WBS feature (the user flagged it twice).
The BnK template is a **live, formula-driven** workbook — we do NOT dump computed
values; we clone the template (keeping its styling + the ``4. Master Data`` ratio
sheet) and rebuild the data sheets so Excel recomputes everything on open:

  * ``2. WBS``       — hierarchy Phase→Module→(SubModule)→Leaf. Only BE + FE/Mobile
                       (and RA for requirement rows) are hard numbers; Ref.Code,
                       Requirement-Analysis, Testing, PM, Total and all roll-ups are
                       formulas. Variable number of rows.
  * ``1. Effort``    — per-module summary; VLOOKUPs the module codes from ``2. WBS``.
  * ``3. Delivery Plan`` — Gantt with **dynamic** month/sprint/week columns sized to
                       the project duration (months=ceil(weeks/4)), so it never runs
                       out of months; plus the 5-milestone table.

Keep the leaf formulas in sync with :mod:`diagram_mcp.wbs_effort` — both encode the
same Master-Data ratio model, one as Excel formulas and one in Python.

The input ``wbs`` dict (see ``wbs_effort`` and the WBS tools) is shaped as::

    {
      "project_info": {"name": str, "project_code": str},
      "phases": [
        {"code": "I", "name": "SET UP & INSTALLATION", "modules": [
          {"code": "I.A", "name": "Solution Design", "groups": [
            {"name": None, "items": [ <leaf>, ... ]},          # name=None → no sub-header
          ]}
        ]}
      ],
      "timeline": {"weeks": int}, "milestones": [...],
    }

where ``<leaf>`` = {name, description, be, fe_mobile, ba, qc, pm, total, remark,
phase_type}. ``be/fe_mobile/ba`` are the only hand numbers; the rest are recomputed
by the formulas (we still store them for the propose summary / validators).
"""

from __future__ import annotations

import math
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.worksheet.worksheet import Worksheet

# The blank BnK template shipped inside the package (cloned for every export).
DEFAULT_TEMPLATE = Path(__file__).resolve().parent / "data" / "wbs_template.xlsx"
# Company logo stamped onto the "0. How to use" cover sheet.
DEFAULT_LOGO = Path(__file__).resolve().parent / "data" / "logo.png"

# ── Column map for the "2. WBS" sheet (1-based) ──────────────────────────────
C_NUM, C_REF, C_FEAT, C_DESC, C_TOTAL = 2, 3, 4, 5, 6
C_BE, C_FEMOB, C_RA, C_TEST, C_PM, C_REMARK = 7, 8, 9, 10, 11, 12
# Optional WBS v2 3-point estimate columns (Optimistic / Most-likely / Pessimistic),
# written only when a project supplies them — left untouched otherwise. They sit AFTER
# the template's columns (B..L) so the Effort sheet's VLOOKUP range ($B$5:$L$) is unaffected.
C_OPT, C_LIK, C_PES = 13, 14, 15
C_P50, C_P80 = 16, 17   # risk-adjusted percentiles (WBS v2)
C_DOD = 18              # Definition-of-Done / acceptance criteria

# Master Data cell refs the template formulas point at.
MD = "'4. Master Data'"
MD_BA, MD_QC, MD_PM = f"{MD}!$C$5", f"{MD}!$C$7", f"{MD}!$C$4"

# Representative template rows to copy styling from (in DATA/[BnK] Template - WBS.xlsx).
_STYLE_ROWS = {"total": 6, "phase": 7, "module": 8, "group": 23, "leaf": 9, "blank": 14}


def _copy_row_style(ws: Worksheet, src_row: int, dst_row: int, cols=range(1, 13)) -> None:
    """Copy cell styling (font/fill/border/alignment/number_format) for a row."""
    for c in cols:
        s = ws.cell(src_row, c)
        d = ws.cell(dst_row, c)
        if s.has_style:
            d._style = copy(s._style)


def _snapshot_styles(ws: Worksheet) -> dict:
    """Grab a deep copy of styling for each row-kind before we overwrite the body."""
    snap: dict = {}
    for kind, r in _STYLE_ROWS.items():
        snap[kind] = {c: copy(ws.cell(r, c)._style) for c in range(1, 13) if ws.cell(r, c).has_style}
    # Row height too (phase/module rows are taller in some templates).
    snap["_heights"] = {kind: ws.row_dimensions[r].height for kind, r in _STYLE_ROWS.items()}
    return snap


def _apply_style(ws: Worksheet, row: int, kind: str, snap: dict) -> None:
    for c, st in snap.get(kind, {}).items():
        ws.cell(row, c)._style = copy(st)
    h = snap.get("_heights", {}).get(kind)
    if h:
        ws.row_dimensions[row].height = h


def _clear_body(ws: Worksheet, first: int, last: int) -> None:
    """Blank out every cell value in the template's example body.

    The template ships footer merges (e.g. B75:K76) inside the body region; any
    that survive turn their non-anchor cells into read-only ``MergedCell`` objects,
    so writing the rebuilt hierarchy there raises "attribute 'value' is read-only".
    Unmerge every range that intersects the clear region first (the title merge in
    rows <first stays intact).
    """
    for mr in list(ws.merged_cells.ranges):
        if mr.max_row >= first and mr.min_row <= last:
            ws.unmerge_cells(str(mr))
    for r in range(first, last + 1):
        for c in range(1, 25):
            cell = ws.cell(r, c)
            if cell.value is not None:
                cell.value = None


def _build_wbs_sheet(ws: Worksheet, wbs: dict, snap: dict) -> dict:
    """Write the Phase→Module→Leaf hierarchy with live formulas. Returns layout
    info (module_rows by code, phase_rows) used by the Effort/Delivery sheets."""
    pi = wbs.get("project_info", {})
    ws.cell(2, C_REF + 1).value = pi.get("project_code", "BNK")  # D2 (col 4) holds Project Code
    title = (pi.get("name") or "OUR SOLUTION").upper()
    ws.cell(3, 2).value = f"WBS OF {title}"

    module_rows: dict[str, int] = {}   # "I.A" -> excel row
    phase_blocks: list[tuple[int, list[int]]] = []  # (phase_row, [module_rows])
    seq = 0
    r = 7  # row 5 header, row 6 grand total, body starts at 7

    # Only surface the 3-point columns when the project actually estimated them, so a
    # plain WBS stays visually identical to the template.
    has_pert = any(
        float(leaf.get("likely") or 0) > 0
        for phase in wbs.get("phases", [])
        for module in phase.get("modules", [])
        for group in module.get("groups", [])
        for leaf in group.get("items", [])
    )
    has_dod = any(
        leaf.get("acceptance_criteria")
        for phase in wbs.get("phases", [])
        for module in phase.get("modules", [])
        for group in module.get("groups", [])
        for leaf in group.get("items", [])
    )
    if has_pert:
        for col, label in ((C_OPT, "Optimistic (O)"), (C_LIK, "Most-likely (M)"),
                           (C_PES, "Pessimistic (P)"), (C_P50, "P50 (md)"), (C_P80, "P80 (md)")):
            ws.cell(5, col).value = label
    if has_dod:
        ws.cell(5, C_DOD).value = "Definition of Done"

    for phase in wbs.get("phases", []):
        phase_row = r
        _apply_style(ws, r, "phase", snap)
        ws.cell(r, 1).value = phase.get("code")
        ws.cell(r, C_NUM).value = phase.get("code")
        ws.cell(r, C_REF).value = phase.get("name")   # non-leaf name lives in col C
        r += 1
        mod_rows_this_phase: list[int] = []

        for module in phase.get("modules", []):
            module_row = r
            module_rows[module.get("code", "")] = module_row
            mod_rows_this_phase.append(module_row)
            _apply_style(ws, r, "module", snap)
            ws.cell(r, 1).value = (module.get("code", "") or "").split(".")[-1]
            ws.cell(r, C_NUM).value = module.get("code")
            ws.cell(r, C_REF).value = module.get("name")   # non-leaf name in col C
            r += 1
            leaf_first = r

            grp_idx = 0
            for group in module.get("groups", []):
                gname = group.get("name")
                if gname:
                    grp_idx += 1
                    _apply_style(ws, r, "group", snap)
                    ws.cell(r, 1).value = grp_idx
                    ws.cell(r, C_REF).value = gname        # non-leaf name in col C
                    r += 1
                for leaf in group.get("items", []):
                    seq += 1
                    _apply_style(ws, r, "leaf", snap)
                    pt = (leaf.get("phase_type") or "development").lower()
                    ws.cell(r, C_NUM).value = seq
                    ws.cell(r, C_REF).value = f'=CONCATENATE($D$2&"-",B{r})'
                    ws.cell(r, C_FEAT).value = leaf.get("name")
                    ws.cell(r, C_DESC).value = leaf.get("description")
                    ws.cell(r, C_TOTAL).value = f"=SUM(G{r}:K{r})"
                    ws.cell(r, C_REMARK).value = leaf.get("remark")
                    _write_leaf_effort(ws, r, leaf, pt)
                    if has_pert and float(leaf.get("likely") or 0) > 0:
                        ws.cell(r, C_OPT).value = leaf.get("optimistic") or None
                        ws.cell(r, C_LIK).value = leaf.get("likely") or None
                        ws.cell(r, C_PES).value = leaf.get("pessimistic") or None
                        ws.cell(r, C_P50).value = leaf.get("pert_p50_md") or None
                        ws.cell(r, C_P80).value = leaf.get("pert_p80_md") or None
                    if has_dod:
                        criteria = leaf.get("acceptance_criteria") or []
                        ws.cell(r, C_DOD).value = "; ".join(criteria) if criteria else None
                    r += 1

            # module roll-up over its full contiguous block (label rows are empty).
            # An empty module (no leaf rows) would make leaf_last < leaf_first, i.e. a
            # reversed range that includes the module row itself → a circular reference
            # that halts Excel's calc chain and blanks every downstream TOTAL. Guard it.
            leaf_last = r - 1
            has_leaves = leaf_last >= leaf_first
            for col, letter in ((C_BE, "G"), (C_FEMOB, "H"), (C_RA, "I"),
                                (C_TEST, "J"), (C_PM, "K")):
                ws.cell(module_row, col).value = (
                    f"=SUM({letter}{leaf_first}:{letter}{leaf_last})" if has_leaves else 0
                )
            ws.cell(module_row, C_TOTAL).value = f"=SUM(G{module_row}:K{module_row})"

            # blank spacer row between modules (styled like template's spacer)
            _apply_style(ws, r, "blank", snap)
            r += 1

        # phase roll-up sums only its module-header rows (avoids the /2 quirk)
        for col, letter in ((C_BE, "G"), (C_FEMOB, "H"), (C_RA, "I"),
                            (C_TEST, "J"), (C_PM, "K")):
            refs = ",".join(f"{letter}{m}" for m in mod_rows_this_phase)
            ws.cell(phase_row, col).value = f"=SUM({refs})" if refs else 0
        ws.cell(phase_row, C_TOTAL).value = f"=SUM(G{phase_row}:K{phase_row})"
        phase_blocks.append((phase_row, mod_rows_this_phase))

    # grand-total row 6 = sum of phase rows
    phase_rows = [pr for pr, _ in phase_blocks]
    for col, letter in ((C_BE, "G"), (C_FEMOB, "H"), (C_RA, "I"),
                        (C_TEST, "J"), (C_PM, "K")):
        refs = ",".join(f"{letter}{p}" for p in phase_rows)
        ws.cell(6, col).value = f"=SUM({refs})" if refs else 0
    ws.cell(6, C_TOTAL).value = "=SUM(G6:K6)"

    return {"module_rows": module_rows, "phase_rows": phase_rows,
            "phase_blocks": phase_blocks, "last_row": r - 1}


def _write_leaf_effort(ws: Worksheet, r: int, leaf: dict, phase_type: str) -> None:
    """Write the hand numbers + ratio formulas for one leaf, gated by phase_type.

    Mirrors the real template: development rows hard-set BE/FE and FORMULA the
    BA/QC/PM; requirement rows hard-set RA; design/uiux/deployment set dev + PM.
    """
    be = float(leaf.get("be", 0) or 0)
    femob = float(leaf.get("fe_mobile", 0) or 0)
    ba = float(leaf.get("ba", 0) or 0)
    dev_range = f"=SUM(G{r}:H{r})"
    pm_formula = f"=SUM(G{r}:J{r})*{MD_PM}"

    if phase_type == "requirement":
        ws.cell(r, C_RA).value = ba or None
        ws.cell(r, C_PM).value = pm_formula
    elif phase_type in ("design", "setup"):
        if be:
            ws.cell(r, C_BE).value = be
        ws.cell(r, C_PM).value = pm_formula
    elif phase_type == "uiux":
        if femob:
            ws.cell(r, C_FEMOB).value = femob
        ws.cell(r, C_PM).value = pm_formula
    elif phase_type == "deployment":
        if be:
            ws.cell(r, C_BE).value = be
        # PM = 0 → leave blank
    else:  # development / support
        if be:
            ws.cell(r, C_BE).value = be
        if femob:
            ws.cell(r, C_FEMOB).value = femob
        ws.cell(r, C_RA).value = f"{dev_range}*{MD_BA}"
        ws.cell(r, C_TEST).value = f"{dev_range}*{MD_QC}"
        ws.cell(r, C_PM).value = pm_formula


def _build_effort_sheet(wb, wbs: dict, wbs_last_row: int) -> None:
    """Rebuild "1. Effort" — one row per phase + module, VLOOKUP'd from "2. WBS".

    The template ships a fixed 9-module list (I..III.B); a real project has a
    different module set, so we regenerate the rows and widen the VLOOKUP range to
    cover the actual WBS body. USD columns keep the template's rate formulas.
    """
    ws = wb["1. Effort"]
    # snapshot styles from the template's phase row (6) and module row (7) + TOTAL (16)
    phase_st = {c: copy(ws.cell(6, c)._style) for c in range(1, 16) if ws.cell(6, c).has_style}
    mod_st = {c: copy(ws.cell(7, c)._style) for c in range(1, 16) if ws.cell(7, c).has_style}
    total_st = {c: copy(ws.cell(16, c)._style) for c in range(1, 16) if ws.cell(16, c).has_style}

    # clear the old body (rows 6..40)
    for r in range(6, 41):
        for c in range(1, 16):
            if ws.cell(r, c).value is not None:
                ws.cell(r, c).value = None

    rng = f"'2. WBS'!$B$5:$L${wbs_last_row}"
    md = "'4. Master Data'"
    r = 6
    phase_rows: list[int] = []
    for phase in wbs.get("phases", []):
        for code, kind in ([(phase.get("code"), "phase")]
                           + [(m.get("code"), "module") for m in phase.get("modules", [])]):
            st = phase_st if kind == "phase" else mod_st
            for c, s in st.items():
                ws.cell(r, c)._style = copy(s)
            ws.cell(r, 2).value = code
            ws.cell(r, 3).value = f"=VLOOKUP($B{r},{rng},2,FALSE)"
            ws.cell(r, 4).value = f"=SUM(E{r}:I{r})"
            for col, idx in ((5, 6), (6, 7), (7, 8), (8, 9), (9, 10)):
                ws.cell(r, col).value = f"=VLOOKUP($B{r},{rng},{idx},FALSE)"
            ws.cell(r, 10).value = f"=SUM(K{r}:O{r})"
            ws.cell(r, 11).value = f"=E{r}*{md}!$C$13"
            ws.cell(r, 12).value = f"=F{r}*{md}!$C$13"
            ws.cell(r, 13).value = f"=G{r}*{md}!$C$12"
            ws.cell(r, 14).value = f"=H{r}*{md}!$C$14"
            ws.cell(r, 15).value = f"=I{r}*{md}!$C$11"
            if kind == "phase":
                phase_rows.append(r)
            r += 1
    # TOTAL row sums the phase rows only
    for c, s in total_st.items():
        ws.cell(r, c)._style = copy(s)
    ws.cell(r, 3).value = "TOTAL"
    refs = ",".join(str(p) for p in phase_rows)
    for col, letter in ((4, "D"), (5, "E"), (6, "F"), (7, "G"), (8, "H"), (9, "I"),
                        (10, "J"), (11, "K"), (12, "L"), (13, "M"), (14, "N"), (15, "O")):
        ws.cell(r, col).value = f"=SUM({','.join(letter+str(p) for p in phase_rows)})" if refs else 0
    # Delete leftover template rows below the TOTAL row.
    if ws.max_row > r:
        ws.delete_rows(r + 1, ws.max_row - r)


_MONTH_ORD = ["1st", "2nd", "3rd", "4th", "5th", "6th", "7th", "8th", "9th",
              "10th", "11th", "12th", "13th", "14th"]
_BNK_MILESTONES = [
    ("Contract signoff", "Contract Signoff"),
    ("Requirement Confirmation/Signoff", "BRD"),
    ("Development Completion and UAT Initiation", "System ready in UAT + Test Cases + Test Report"),
    ("Completion of UAT", "Source code + User Guide + Technical Specification"),
    ("Completion of Post-Launch Support", "Final delivery package + maintenance log"),
]
_DEFAULT_RESOURCES = ["Project Manager", "Technical Lead", "Developer",
                      "Business Analyst", "Quality Controller", "Designer", "Devops"]


def _module_schedule(wbs: dict, weeks: int) -> list[dict]:
    """Allocate a waterfall week-range to each module proportional to its effort.

    Returns [{code, name, start_week, end_week}] with end_week<=weeks so the Gantt
    never overflows the calendar (the "not enough months" failure the user flagged).
    Honours an explicit per-module ``{"start_week","end_week"}`` if the planner set one.
    """
    mods = []
    for phase in wbs.get("phases", []):
        for m in phase.get("modules", []):
            eff = 0.0
            for g in m.get("groups", []):
                for it in g.get("items", []):
                    eff += float(it.get("total", 0) or 0)
            mods.append({"code": m.get("code"), "name": m.get("name"),
                         "effort": eff, "start_week": m.get("start_week"),
                         "end_week": m.get("end_week")})
    total = sum(m["effort"] for m in mods) or 1.0
    cum = 0.0
    for m in mods:
        if m["start_week"] and m["end_week"]:
            m["start_week"] = max(1, min(weeks, int(m["start_week"])))
            m["end_week"] = max(m["start_week"], min(weeks, int(m["end_week"])))
            continue
        start = int(round(weeks * cum / total)) + 1
        cum += m["effort"]
        end = int(round(weeks * cum / total))
        m["start_week"] = max(1, min(weeks, start))
        m["end_week"] = max(m["start_week"], min(weeks, end))
    return mods


def _build_delivery_sheet(wb, wbs: dict, wbs_last_row: int) -> dict:
    """Rebuild "3. Delivery Plan" with a DYNAMIC month/sprint/week grid.

    months=ceil(weeks/4), 2 sprints/month, 1 sprint=2 weeks. Header columns,
    merges and the per-module Gantt bars all scale to the project duration, so the
    plan always shows enough months (the template's fixed 20-week grid does not).
    """
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from wbs_effort import delivery_grid

    ws = wb["3. Delivery Plan"]
    weeks = int(wbs.get("timeline", {}).get("weeks") or 16)
    grid = delivery_grid(weeks)
    weeks = grid["weeks"]

    # clear a generous region (old fixed grid + body)
    for mr in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(mr))
    for r in range(1, 70):
        for c in range(1, 45):
            if ws.cell(r, c).value is not None:
                ws.cell(r, c).value = None
            ws.cell(r, c).fill = PatternFill(fill_type=None)

    thin = Side(style="thin", color="FFBFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    month_fill = PatternFill("solid", fgColor="FFD9E1F2")
    sprint_fill = PatternFill("solid", fgColor="FFE2EFDA")
    head_fill = PatternFill("solid", fgColor="FF0094AA")
    bar_fill = PatternFill("solid", fgColor="FF4DB6AC")
    center = Alignment(horizontal="center", vertical="center")
    bold = Font(bold=True)
    whitebold = Font(bold=True, color="FFFFFFFF")

    FIRST = 6  # first week column (F)
    ws.cell(1, 2).value = "Master Plan"; ws.cell(1, 2).font = Font(bold=True, size=14)

    # row 3 = months (merge 4 cols), row 4 = sprints (merge 2 cols), row 5 = weeks
    for mi in range(grid["months"]):
        c0 = FIRST + mi * 4
        ws.merge_cells(start_row=3, start_column=c0, end_row=3, end_column=c0 + 3)
        mc = ws.cell(3, c0)
        mc.value = f"{_MONTH_ORD[mi] if mi < len(_MONTH_ORD) else str(mi+1)+'th'} Month"
        mc.fill = month_fill; mc.font = bold; mc.alignment = center
    for si in range(grid["sprints"]):
        c0 = FIRST + si * 2
        ws.merge_cells(start_row=4, start_column=c0, end_row=4, end_column=c0 + 1)
        sc = ws.cell(4, c0)
        sc.value = f"Sprint {si+1}"; sc.fill = sprint_fill; sc.font = bold; sc.alignment = center
    for hc, lbl in ((2, "#"), (3, "Module"), (4, "Start"), (5, "End")):
        cell = ws.cell(5, hc); cell.value = lbl; cell.fill = head_fill
        cell.font = whitebold; cell.alignment = center; cell.border = border
    for wi in range(weeks):
        cell = ws.cell(5, FIRST + wi)
        cell.value = f"W{wi+1}"; cell.fill = head_fill; cell.font = whitebold
        cell.alignment = center; cell.border = border
        ws.column_dimensions[get_column_letter(FIRST + wi)].width = 4.5

    # module rows with Gantt bars
    rng = f"'2. WBS'!$B$5:$L${wbs_last_row}"
    sched = _module_schedule(wbs, weeks)
    r = 6
    for i, m in enumerate(sched, 1):
        ws.cell(r, 2).value = m["code"]
        ws.cell(r, 3).value = f"=VLOOKUP($B{r},{rng},2,FALSE)"
        ws.cell(r, 4).value = f"W{m['start_week']}"
        ws.cell(r, 5).value = f"W{m['end_week']}"
        for c in range(2, 6):
            ws.cell(r, c).border = border
        for wk in range(m["start_week"], m["end_week"] + 1):
            bar = ws.cell(r, FIRST + wk - 1)
            bar.fill = bar_fill; bar.border = border
        r += 1
    last_module_row = r - 1

    # resource planning
    r += 1
    ws.cell(r, 3).value = "Resource Planning"; ws.cell(r, 3).font = bold
    r += 1
    for i, role in enumerate(_DEFAULT_RESOURCES, 1):
        ws.cell(r, 2).value = i
        ws.cell(r, 3).value = role
        ws.cell(r, 4).value = "TBD"; ws.cell(r, 5).value = "TBD"
        r += 1

    # milestones
    # The "Deliverables" text is free-form prose (e.g. "System ready in UAT +
    # Test Cases + Test Report"), but column F alone is only 4.5 wide (sized for
    # a Gantt week column) — wrapped + vertically-centered text there renders as
    # an unreadable single letter per line. Merge it across a wide span (mirroring
    # the template's F:U milestone merges) so the sentence actually fits.
    DELIV_LAST = FIRST + max(weeks, 16) - 1
    r += 1
    hdr = ["#", "Deliverables Milestone", "Start", "End", "Deliverables"]
    for c, lbl in zip(range(2, 7), hdr):
        cell = ws.cell(r, c); cell.value = lbl; cell.fill = head_fill
        cell.font = whitebold; cell.border = border
    ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=DELIV_LAST)
    ws.cell(r, 6).alignment = center
    r += 1
    milestones = wbs.get("milestones") or [
        {"name": n, "deliverables": [d]} for n, d in _BNK_MILESTONES
    ]
    for i, ms in enumerate(milestones, 1):
        ws.cell(r, 2).value = i
        ws.cell(r, 3).value = ms.get("name")
        ws.cell(r, 4).value = ms.get("start", "TBD")
        ws.cell(r, 5).value = ms.get("end", "TBD")
        dl = ms.get("deliverables")
        deliv_cell = ws.cell(r, 6)
        deliv_cell.value = ", ".join(dl) if isinstance(dl, list) else (dl or "")
        deliv_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=DELIV_LAST)
        for c in range(2, DELIV_LAST + 1):
            ws.cell(r, c).border = border
        ws.row_dimensions[r].height = 30
        r += 1

    return {"weeks": weeks, "months": grid["months"], "sprints": grid["sprints"],
            "module_rows": last_module_row - 5, "schedule": sched}


def _write_master_data_ratios(wb, ratios: dict) -> None:
    """Sync the ``4. Master Data`` ratio cells with the ratios used for the summary.

    The template ships with PM=0.05; if we leave it, Excel recomputes PM with 0.05
    while our Python summary/validators use a different ratio → silent mismatch.
    Writing the cells keeps the live workbook and the propose summary consistent.
    """
    md = wb["4. Master Data"]
    md["C4"].value = float(ratios.get("pm_on_total", 0.10))   # PM, on (dev+ba+qc)
    md["C5"].value = float(ratios.get("ba_on_dev", 0.10))     # BA, on dev
    md["C7"].value = float(ratios.get("qc_on_dev", 0.30))     # QC, on dev


def _write_master_data_rate_card(wb, rate_card: dict[str, float] | None = None) -> None:
    """Fill the ``4. Master Data`` Rate column (C11:C14) — ships empty in the template.

    ``1. Effort`` col K/L/M/N/O multiply MD directly against these cells
    (``K6 = E6 * '4. Master Data'!$C$13``, etc.) — so these MUST be a per-MAN-DAY USD
    rate, not a monthly rate. ``rate_card`` is USD/month (the BnK-quoted unit); convert
    with :func:`wbs_effort.rate_per_manday` (20-workday month, NOT the 22 used for
    total_manmonths elsewhere — see that module for why).
    """
    from wbs_effort import DEFAULT_RATE_CARD_USD_PER_MONTH, rate_per_manday

    rc = rate_card or DEFAULT_RATE_CARD_USD_PER_MONTH
    md = wb["4. Master Data"]
    md["C11"].value = rate_per_manday(rc.get("PM", 0))          # PM
    md["C12"].value = rate_per_manday(rc.get("BA", 0))           # BA
    md["C13"].value = rate_per_manday(rc.get("Developer", 0))   # Developer (BE + FE/Mobile)
    md["C14"].value = rate_per_manday(rc.get("QC", 0))           # QC


# The 5 canonical sheets the deliverable keeps; the template also ships auxiliary
# reference sheets (General_BK / By-Month / Summary) whose VLOOKUPs would go stale
# once we rebuild the WBS, so we drop them for a clean client file.
_KEEP_SHEETS = ["0. How to use", "1. Effort", "2. WBS", "3. Delivery Plan", "4. Master Data"]


def _add_logo(wb, logo_path: str | Path | None = None) -> None:
    """Stamp the company logo onto the top-left of the "0. How to use" sheet."""
    path = Path(logo_path or DEFAULT_LOGO)
    if "0. How to use" not in wb.sheetnames or not path.exists():
        return
    img = XLImage(str(path))
    # Keep the native aspect ratio but cap the footprint on the cover sheet.
    max_side = 110
    if img.width and img.height:
        scale = min(max_side / img.width, max_side / img.height, 1.0)
        img.width = int(img.width * scale)
        img.height = int(img.height * scale)
    wb["0. How to use"].add_image(img, "A1")


def build_wbs_workbook(wbs: dict, out_path: str | Path,
                       template_path: str | Path | None = None) -> dict:
    """Clone the template and fill it from ``wbs``. Returns layout metadata."""
    wb = load_workbook(template_path or DEFAULT_TEMPLATE, data_only=False)
    for name in list(wb.sheetnames):
        if name not in _KEEP_SHEETS:
            del wb[name]
    _write_master_data_ratios(wb, wbs.get("ratios", {}))
    _add_logo(wb)
    ws = wb["2. WBS"]
    snap = _snapshot_styles(ws)
    # Old example body spans rows 6..~73; clear generously.
    _clear_body(ws, 6, max(80, ws.max_row))
    layout = _build_wbs_sheet(ws, wbs, snap)
    _build_effort_sheet(wb, wbs, layout["last_row"])
    layout["delivery"] = _build_delivery_sheet(wb, wbs, layout["last_row"])

    # Delete leftover template rows after the last data row so the sheet ends
    # exactly where the data ends (no trailing empty/styled rows from the template).
    last_wbs = layout["last_row"]
    if ws.max_row > last_wbs:
        ws.delete_rows(last_wbs + 1, ws.max_row - last_wbs)

    wb.save(out_path)
    return layout
